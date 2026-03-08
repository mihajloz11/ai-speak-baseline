from __future__ import annotations

from dataclasses import asdict, dataclass
from pathlib import Path
import json
import zipfile

import librosa
import numpy as np
import pandas as pd
import soundfile as sf
from openpyxl import load_workbook
from tqdm import tqdm

from .constants import BLENDSHAPE_DIM, DEFAULT_AUDIO_SAMPLE_RATE, DEFAULT_FPS


@dataclass
class FeatureConfig:
    sample_rate: int = DEFAULT_AUDIO_SAMPLE_RATE
    fps: int = DEFAULT_FPS
    n_mels: int = 80
    n_fft: int = 1024
    hop_length: int = 160
    win_length: int = 400


def ensure_directory(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def safe_unzip(zip_path: Path, destination: Path, force: bool = False) -> Path:
    destination = ensure_directory(destination)
    if not force and any(destination.iterdir()):
        return destination

    with zipfile.ZipFile(zip_path, "r") as archive:
        destination_root = destination.resolve()
        for member in archive.infolist():
            member_path = destination / member.filename
            resolved = member_path.resolve()
            if not str(resolved).startswith(str(destination_root)):
                raise ValueError(f"Unsafe zip entry detected: {member.filename}")
            archive.extract(member, destination)
    return destination


def parse_transcript_workbook(workbook_path: Path) -> dict[int, str]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook.active
    transcripts: dict[int, str] = {}
    for row in sheet.iter_rows(values_only=True):
        if not row or row[0] is None or row[1] is None:
            continue
        try:
            sample_index = int(str(row[0]).strip())
        except ValueError:
            continue
        transcripts[sample_index] = str(row[1]).strip()
    workbook.close()
    return transcripts


def read_blendshape_csv(path: Path) -> np.ndarray:
    values = np.loadtxt(path, delimiter=",", dtype=np.float32)
    if values.ndim == 1:
        values = values[None, :]
    if values.shape[1] != BLENDSHAPE_DIM:
        raise ValueError(f"{path} has {values.shape[1]} columns, expected {BLENDSHAPE_DIM}.")
    return values


def load_audio(path: Path, sample_rate: int) -> np.ndarray:
    audio, _ = librosa.load(path.as_posix(), sr=sample_rate, mono=True)
    return audio.astype(np.float32)


def _safe_delta(features: np.ndarray, order: int) -> np.ndarray:
    if features.shape[1] < 3:
        return np.zeros_like(features)
    width = min(9, features.shape[1])
    if width % 2 == 0:
        width -= 1
    if width < 3:
        return np.zeros_like(features)
    return librosa.feature.delta(features, order=order, width=width)


def interpolate_sequence(values: np.ndarray, target_frames: int) -> np.ndarray:
    if target_frames <= 0:
        raise ValueError("target_frames must be positive.")
    if values.shape[0] == target_frames:
        return values.astype(np.float32)
    if values.shape[0] == 1:
        return np.repeat(values.astype(np.float32), target_frames, axis=0)

    source_x = np.linspace(0.0, 1.0, num=values.shape[0], dtype=np.float32)
    target_x = np.linspace(0.0, 1.0, num=target_frames, dtype=np.float32)
    interpolated = np.stack(
        [np.interp(target_x, source_x, values[:, index]) for index in range(values.shape[1])],
        axis=1,
    )
    return interpolated.astype(np.float32)


def extract_frame_features(audio_path: Path, target_frames: int, config: FeatureConfig) -> np.ndarray:
    audio = load_audio(audio_path, config.sample_rate)
    mel = librosa.feature.melspectrogram(
        y=audio,
        sr=config.sample_rate,
        n_fft=config.n_fft,
        hop_length=config.hop_length,
        win_length=config.win_length,
        n_mels=config.n_mels,
        power=2.0,
    )
    log_mel = librosa.power_to_db(mel, ref=np.max).astype(np.float32)
    delta = _safe_delta(log_mel, order=1).astype(np.float32)
    delta2 = _safe_delta(log_mel, order=2).astype(np.float32)
    rms = librosa.feature.rms(y=audio, frame_length=config.win_length, hop_length=config.hop_length).astype(
        np.float32
    )
    feature_frames = np.concatenate([log_mel, delta, delta2, rms], axis=0).T
    return interpolate_sequence(feature_frames, target_frames)


def build_text_vocab(texts: list[str]) -> dict[str, int]:
    charset = sorted({character for text in texts for character in text.lower()})
    vocab = {"<pad>": 0, "<unk>": 1}
    for character in charset:
        if character not in vocab:
            vocab[character] = len(vocab)
    return vocab


def encode_text(text: str, vocab: dict[str, int]) -> list[int]:
    if not text:
        return []
    return [vocab.get(character, vocab["<unk>"]) for character in text.lower()]


def build_manifest(extracted_root: Path) -> pd.DataFrame:
    labels_root = extracted_root / "labels_aligned" / "labels_aligned" / "per_phoneme"
    records: list[dict[str, object]] = []
    for speaker in ("spk08", "spk14"):
        speaker_root = extracted_root / f"{speaker}_blendshapes"
        audio_root = speaker_root / f"renamed_{speaker}"
        transcript_workbook = next(speaker_root.glob("*transcript*.xlsx"))
        transcripts = parse_transcript_workbook(transcript_workbook)
        for audio_path in sorted(audio_root.glob("*.wav")):
            sample_id = audio_path.stem
            target_path = audio_path.with_suffix(".csv")
            target = read_blendshape_csv(target_path)
            info = sf.info(audio_path.as_posix())
            sample_index = int(sample_id.split("_")[1])
            phoneme_path = labels_root / f"{sample_id}.txt"
            records.append(
                {
                    "sample_id": sample_id,
                    "speaker": speaker,
                    "sample_index": sample_index,
                    "audio_path": audio_path.as_posix(),
                    "target_path": target_path.as_posix(),
                    "phoneme_path": phoneme_path.as_posix() if phoneme_path.exists() else "",
                    "transcript": transcripts.get(sample_index, ""),
                    "num_frames": int(target.shape[0]),
                    "duration_sec": float(info.frames / info.samplerate) if info.samplerate else 0.0,
                }
            )
    manifest = pd.DataFrame.from_records(records).sort_values(["speaker", "sample_index"]).reset_index(drop=True)
    return manifest


def precompute_features(
    manifest: pd.DataFrame,
    features_root: Path,
    config: FeatureConfig,
    force: bool = False,
) -> pd.DataFrame:
    features_root = ensure_directory(features_root)
    feature_paths: list[str] = []
    for row in tqdm(manifest.itertuples(index=False), total=len(manifest), desc="Caching features"):
        feature_path = features_root / f"{row.sample_id}.npy"
        if force or not feature_path.exists():
            features = extract_frame_features(Path(row.audio_path), int(row.num_frames), config)
            np.save(feature_path, features)
        feature_paths.append(feature_path.as_posix())

    updated = manifest.copy()
    updated["feature_path"] = feature_paths
    return updated


def split_manifest(manifest: pd.DataFrame, validation_ratio: float = 0.15, seed: int = 42) -> tuple[pd.DataFrame, pd.DataFrame]:
    rng = np.random.default_rng(seed)
    train_parts: list[pd.DataFrame] = []
    validation_parts: list[pd.DataFrame] = []
    for _, speaker_rows in manifest.groupby("speaker", sort=True):
        order = rng.permutation(len(speaker_rows))
        validation_size = max(1, int(round(len(speaker_rows) * validation_ratio)))
        validation_indices = set(speaker_rows.iloc[order[:validation_size]].index.tolist())
        validation_parts.append(speaker_rows.loc[sorted(validation_indices)])
        train_parts.append(speaker_rows.drop(index=validation_indices))
    train_df = pd.concat(train_parts, axis=0).sort_values(["speaker", "sample_index"]).reset_index(drop=True)
    val_df = pd.concat(validation_parts, axis=0).sort_values(["speaker", "sample_index"]).reset_index(drop=True)
    return train_df, val_df


def save_json(path: Path, payload: dict) -> None:
    ensure_directory(path.parent)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2, ensure_ascii=False)


def save_dataset_info(
    path: Path,
    manifest: pd.DataFrame,
    feature_config: FeatureConfig,
    feature_dim: int,
    vocab: dict[str, int],
) -> None:
    payload = {
        "num_samples": int(len(manifest)),
        "speakers": sorted(manifest["speaker"].unique().tolist()),
        "feature_dim": int(feature_dim),
        "feature_config": asdict(feature_config),
        "text_vocab": vocab,
    }
    save_json(path, payload)


